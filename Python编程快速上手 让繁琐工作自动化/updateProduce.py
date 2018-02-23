# python3
# updateProduce.py 更新价格

import openpyxl

#更新的产品和他们的价格
PRICE_UPDATES = {'Grlic' : 999999.07,
                 'Celery' : 999999.19,
                 'Lemon' : 999999.27}
#打开excel
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']
#循环行并且更新价格
for rowNum in range(2, sheet.max_row):#跳过第一行
    produceName = sheet.cell(row = rowNum, column = 1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row = rowNum, column = 2).value = PRICE_UPDATES[produceName]

wb.save('updateProduceSales.xlsx') 

